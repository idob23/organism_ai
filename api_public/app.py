"""Deduplication API v1 — standalone FastAPI service.

Usage:
    uvicorn api_public.app:app --host 0.0.0.0 --port 8080
"""

import csv
import io
import os
import tempfile
import time
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Any

from fastapi import FastAPI, Form, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
import structlog
from dotenv import load_dotenv

load_dotenv()

from auth import get_max_entities, get_tier, validate_key
from dedup import find_duplicates
from rate_limit import (
    check_ip_rate_limit,
    check_rate_limit,
    get_usage_today,
    record_ip_request,
    record_request,
)
from usage import get_usage_stats, record_usage_background

structlog.configure(
    processors=[
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.JSONRenderer(),
    ],
)
_log = structlog.get_logger("app")

API_VERSION = "1.0.0"

app = FastAPI(
    title="Deduplication API",
    description="Find duplicate entities using semantic similarity (embeddings + cosine).",
    version=API_VERSION,
    docs_url="/docs",
    redoc_url="/redoc",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Request / Response models ──────────────────────────────────────

class DeduplicateRequest(BaseModel):
    entities: list[str] = Field(
        ..., min_length=2, max_length=500,
        description="List of entity names to check for duplicates (2\u2013500).",
    )
    threshold: float = Field(
        default=0.85, ge=0.5, le=1.0,
        description="Similarity threshold (0.5\u20131.0). Higher = stricter.",
    )


class DuplicateGroupResponse(BaseModel):
    items: list[str]
    similarity: float


class DeduplicateResponse(BaseModel):
    groups: list[DuplicateGroupResponse]
    total_entities: int
    duplicates_found: int
    processing_time_ms: int


class HealthResponse(BaseModel):
    status: str
    version: str


class UsageResponse(BaseModel):
    requests_today: int
    requests_this_month: int
    plan: str


class ErrorResponse(BaseModel):
    detail: str


# ── Auth dependency ────────────────────────────────────────────────

def _require_api_key(x_api_key: str | None = Header(None)) -> str:
    if not x_api_key:
        raise HTTPException(status_code=401, detail="Missing X-API-Key header")
    if not validate_key(x_api_key):
        raise HTTPException(status_code=401, detail="Invalid API key")
    return x_api_key


# ── Endpoints ──────────────────────────────────────────────────────

@app.get("/v1/health", response_model=HealthResponse)
async def health():
    return HealthResponse(status="ok", version=API_VERSION)


@app.post(
    "/v1/deduplicate",
    response_model=DeduplicateResponse,
    responses={
        401: {"model": ErrorResponse},
        422: {"model": ErrorResponse},
        429: {"model": ErrorResponse},
        500: {"model": ErrorResponse},
    },
)
async def deduplicate(
    body: DeduplicateRequest,
    x_api_key: str | None = Header(None),
):
    api_key = _require_api_key(x_api_key)

    # Rate limit check
    allowed, remaining = check_rate_limit(api_key)
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail="Rate limit exceeded. Upgrade your plan for higher limits.",
        )

    # Entity count limit per tier
    max_ent = get_max_entities(api_key)
    if len(body.entities) > max_ent:
        raise HTTPException(
            status_code=422,
            detail=f"Too many entities: {len(body.entities)}. "
                   f"Your plan allows max {max_ent} per request.",
        )

    # Record rate limit hit
    record_request(api_key)

    try:
        result = await find_duplicates(body.entities, body.threshold)
    except Exception as e:
        _log.error("deduplicate_failed", error=str(e))
        raise HTTPException(status_code=500, detail="Internal processing error")

    groups_resp = [
        DuplicateGroupResponse(items=g.items, similarity=g.similarity)
        for g in result.groups
    ]

    # Fire-and-forget usage tracking
    record_usage_background(
        api_key=api_key,
        endpoint="/v1/deduplicate",
        entities_count=result.total_entities,
        groups_found=len(result.groups),
        processing_time_ms=result.processing_time_ms,
    )

    return DeduplicateResponse(
        groups=groups_resp,
        total_entities=result.total_entities,
        duplicates_found=result.duplicates_found,
        processing_time_ms=result.processing_time_ms,
    )


@app.get(
    "/v1/usage",
    response_model=UsageResponse,
    responses={401: {"model": ErrorResponse}},
)
async def usage(x_api_key: str | None = Header(None)):
    api_key = _require_api_key(x_api_key)
    stats = await get_usage_stats(api_key)
    tier = get_tier(api_key)
    return UsageResponse(
        requests_today=stats["requests_today"],
        requests_this_month=stats["requests_this_month"],
        plan=tier,
    )


# ── Web UI: file upload dedup ─────────────────────────────────────

# In-memory session storage for download reports
# {session_id: (timestamp, DeduplicationResult, original_values)}
_sessions: dict[str, tuple[float, Any, list[str]]] = {}
SESSION_TTL = 1800  # 30 minutes

MAX_COLUMN_VALUES = 500
MAX_UPLOAD_SIZE = 5 * 1024 * 1024  # 5 MB


def _cleanup_sessions() -> None:
    """Remove expired sessions."""
    now = time.time()
    expired = [k for k, (ts, _, _) in _sessions.items() if now - ts > SESSION_TTL]
    for k in expired:
        del _sessions[k]


def fix_1c_xlsx(input_bytes: bytes) -> bytes:
    """Fix 1C SharedStrings.xml capitalization for openpyxl compatibility."""
    tmp_in = tempfile.mktemp(suffix=".xlsx")
    tmp_out = tempfile.mktemp(suffix=".xlsx")
    try:
        with open(tmp_in, "wb") as f:
            f.write(input_bytes)
        with zipfile.ZipFile(tmp_in, "r") as zin:
            with zipfile.ZipFile(tmp_out, "w") as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    name = item.filename
                    if name == "xl/SharedStrings.xml":
                        name = "xl/sharedStrings.xml"
                    zout.writestr(name, data)
        with open(tmp_out, "rb") as f:
            return f.read()
    finally:
        for p in (tmp_in, tmp_out):
            if os.path.exists(p):
                os.unlink(p)


def _parse_xlsx(file_bytes: bytes, column_name: str) -> list[str]:
    """Parse xlsx file and extract values from the specified column."""
    import openpyxl

    fixed = fix_1c_xlsx(file_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(fixed), read_only=True, data_only=True)
    ws = wb.active

    # Find column index by header name
    headers = []
    col_idx = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip() if c is not None else "" for c in row]
        break

    for i, h in enumerate(headers):
        if h == column_name:
            col_idx = i
            break

    if col_idx is None:
        wb.close()
        raise ValueError(
            f"Column '{column_name}' not found. "
            f"Available: {', '.join(h for h in headers if h)}"
        )

    values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(values) >= MAX_COLUMN_VALUES:
            break
        if col_idx < len(row) and row[col_idx] is not None:
            val = str(row[col_idx]).strip()
            if val:
                values.append(val)

    wb.close()
    return values


def _parse_csv(file_bytes: bytes, column_name: str) -> list[str]:
    """Parse csv file and extract values from the specified column."""
    text = file_bytes.decode("utf-8-sig", errors="replace")

    # Auto-detect delimiter
    try:
        dialect = csv.Sniffer().sniff(text[:8192])
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    headers = next(reader, None)
    if not headers:
        raise ValueError("CSV file is empty")

    headers = [h.strip() for h in headers]
    col_idx = None
    for i, h in enumerate(headers):
        if h == column_name:
            col_idx = i
            break

    if col_idx is None:
        raise ValueError(
            f"Column '{column_name}' not found. "
            f"Available: {', '.join(h for h in headers if h)}"
        )

    values = []
    for row in reader:
        if len(values) >= MAX_COLUMN_VALUES:
            break
        if col_idx < len(row):
            val = row[col_idx].strip()
            if val:
                values.append(val)

    return values


class FileDeduplicateResponse(BaseModel):
    session_id: str
    groups: list[DuplicateGroupResponse]
    total_entities: int
    duplicates_found: int
    processing_time_ms: int


@app.post(
    "/v1/deduplicate-file",
    response_model=FileDeduplicateResponse,
    responses={
        422: {"model": ErrorResponse},
        429: {"model": ErrorResponse},
        500: {"model": ErrorResponse},
    },
)
async def deduplicate_file(
    request: Request,
    file: UploadFile,
    column_name: str = Form(...),
    threshold: float = Form(default=0.92),
):
    _cleanup_sessions()

    # IP rate limit
    client_ip = request.headers.get("X-Real-IP") \
        or request.headers.get("X-Forwarded-For", "").split(",")[0].strip() \
        or (request.client.host if request.client else "unknown")
    allowed, remaining = check_ip_rate_limit(client_ip)
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail="\u041b\u0438\u043c\u0438\u0442 \u0437\u0430\u043f\u0440\u043e\u0441\u043e\u0432 "
                   "\u0438\u0441\u0447\u0435\u0440\u043f\u0430\u043d (5 \u0432 \u0434\u0435\u043d\u044c). "
                   "\u041f\u043e\u043f\u0440\u043e\u0431\u0443\u0439\u0442\u0435 \u0437\u0430\u0432\u0442\u0440\u0430 "
                   "\u0438\u043b\u0438 \u0441\u0432\u044f\u0436\u0438\u0442\u0435\u0441\u044c: @organism_ai",
        )

    # Read file
    file_bytes = await file.read()
    if len(file_bytes) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=422,
            detail="\u0424\u0430\u0439\u043b \u0441\u043b\u0438\u0448\u043a\u043e\u043c "
                   "\u0431\u043e\u043b\u044c\u0448\u043e\u0439 (\u043c\u0430\u043a\u0441 5 \u041c\u0411)",
        )

    # Determine format and parse
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=422,
            detail="Unsupported file format. Use .xlsx, .xls, or .csv",
        )

    try:
        if fname.endswith(".csv"):
            values = _parse_csv(file_bytes, column_name)
        else:
            values = _parse_xlsx(file_bytes, column_name)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        _log.error("file_parse_failed", error=str(e), filename=fname)
        raise HTTPException(
            status_code=422,
            detail="\u041e\u0448\u0438\u0431\u043a\u0430 \u0447\u0442\u0435\u043d\u0438\u044f "
                   "\u0444\u0430\u0439\u043b\u0430: " + str(e),
        )

    if len(values) < 2:
        raise HTTPException(
            status_code=422,
            detail="\u041d\u0435\u0434\u043e\u0441\u0442\u0430\u0442\u043e\u0447\u043d\u043e "
                   "\u0437\u0430\u043f\u0438\u0441\u0435\u0439 \u0434\u043b\u044f \u043f\u043e\u0438\u0441\u043a\u0430 "
                   "\u0434\u0443\u0431\u043b\u0438\u043a\u0430\u0442\u043e\u0432 (\u043c\u0438\u043d\u0438\u043c\u0443\u043c 2)",
        )

    record_ip_request(client_ip)

    try:
        result = await find_duplicates(values, threshold)
    except Exception as e:
        _log.error("file_deduplicate_failed", error=str(e))
        raise HTTPException(status_code=500, detail="Internal processing error")

    # Store session for report download
    session_id = str(uuid.uuid4())
    _sessions[session_id] = (time.time(), result, values)

    groups_resp = [
        DuplicateGroupResponse(items=g.items, similarity=g.similarity)
        for g in result.groups
    ]

    return FileDeduplicateResponse(
        session_id=session_id,
        groups=groups_resp,
        total_entities=result.total_entities,
        duplicates_found=result.duplicates_found,
        processing_time_ms=result.processing_time_ms,
    )


@app.get("/v1/download-report")
async def download_report(session_id: str):
    _cleanup_sessions()

    if session_id not in _sessions:
        raise HTTPException(status_code=404, detail="Session not found or expired")

    ts, result, values = _sessions[session_id]
    if time.time() - ts > SESSION_TTL:
        del _sessions[session_id]
        raise HTTPException(status_code=404, detail="Session expired")

    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()

    # Sheet 1: Duplicates
    ws1 = wb.active
    ws1.title = "\u0414\u0443\u0431\u043b\u0438\u043a\u0430\u0442\u044b"

    # Headers
    headers = [
        "\u0413\u0440\u0443\u043f\u043f\u0430",
        "\u0417\u043d\u0430\u0447\u0435\u043d\u0438\u0435",
        "\u0421\u0445\u043e\u0434\u0441\u0442\u0432\u043e %",
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for i, group in enumerate(result.groups, 1):
        sim_pct = round(group.similarity * 100, 1)
        for j, item in enumerate(group.items):
            ws1.cell(row=row_num, column=1, value=i if j == 0 else "")
            ws1.cell(row=row_num, column=2, value=item)
            ws1.cell(row=row_num, column=3, value=sim_pct if j == 0 else "")
            row_num += 1

    # Auto-width columns
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 15

    # Sheet 2: Summary
    ws2 = wb.create_sheet(title="\u0421\u0432\u043e\u0434\u043a\u0430")
    summary_data = [
        ("\u0412\u0441\u0435\u0433\u043e \u0437\u0430\u043f\u0438\u0441\u0435\u0439", result.total_entities),
        ("\u0413\u0440\u0443\u043f\u043f \u0434\u0443\u0431\u043b\u0438\u043a\u0430\u0442\u043e\u0432", len(result.groups)),
        (
            "\u0417\u0430\u043f\u0438\u0441\u0435\u0439-\u0434\u0443\u0431\u043b\u0438\u043a\u0430\u0442\u043e\u0432",
            result.duplicates_found,
        ),
        (
            "\u0412\u0440\u0435\u043c\u044f \u043e\u0431\u0440\u0430\u0431\u043e\u0442\u043a\u0438 (\u043c\u0441)",
            result.processing_time_ms,
        ),
        (
            "\u0414\u0430\u0442\u0430 \u0430\u043d\u0430\u043b\u0438\u0437\u0430",
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        ),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_label = ws2.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=value)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=dedup_report.xlsx"
        },
    )


# ── Root: serve web UI ────────────────────────────────────────────

@app.get("/")
async def root():
    return FileResponse(
        os.path.join(os.path.dirname(__file__), "static", "index.html")
    )


# ── Static files (MUST be after all route definitions) ────────────

_static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(_static_dir):
    app.mount("/static", StaticFiles(directory=_static_dir), name="static")
